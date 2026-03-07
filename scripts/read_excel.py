import openpyxl, sys

def dump_sheet(path, max_rows=50):
    wb = openpyxl.load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"=== Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column}) ===")
        for r in range(1, min(ws.max_row + 1, max_rows + 1)):
            row = []
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(r, c).value
                if v is not None:
                    row.append(f"C{c}={v}")
            if row:
                sep = "  |  "
                print(f"  R{r}: {sep.join(row)}")

if __name__ == "__main__":
    dump_sheet(sys.argv[1], int(sys.argv[2]) if len(sys.argv) > 2 else 50)
